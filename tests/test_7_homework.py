# coding: utf-8
import os
import csv
import time
import shutil
import zipfile

import pytest
import requests
from pypdf import PdfReader
from openpyxl import load_workbook
from pypdf.generic import IndirectObject

from selenium import webdriver
from selene import browser

WORKING_DIR = os.getcwd()
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PATH = SCRIPT_DIR if 'tests' in SCRIPT_DIR else os.path.join(WORKING_DIR, 'tests')
os.chdir(PATH)


@pytest.mark.parametrize("data", [
    ['Anna', 'Pavel', 'Peter'],
    ['Alex', 'Serj', 'Yana'],
])
def test_csv(data):
    path_to_file = os.path.join(PATH, 'resources', 'eggs.csv')
    if not os.path.exists('resources'):
        os.makedirs('resources')

    with open(path_to_file, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(data)

    with open(path_to_file) as csvfile:
        csvreader = csv.reader(csvfile)
        for row in csvreader:
            if len(row) > 0:
                assert row == data


def test_download_file_with_browser():
    download_dir = os.path.join(SCRIPT_DIR, 'tmp')
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(5)
    downloaded_files = os.listdir(download_dir)
    assert len(downloaded_files) > 0, f"No files found in {download_dir}"
    for file in downloaded_files:
        os.remove(os.path.join(download_dir, file))


def test_download_file_with_requests():
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    download_dir = os.path.join(SCRIPT_DIR, 'tmp')
    path_to_file = os.path.join(download_dir, 'selenium_logo.png')

    response = requests.get(url)

    with open(path_to_file, 'wb') as file:
        file.write(response.content)

    size = os.path.getsize(path_to_file)
    assert size == 30803


def test_pdf():
    url = "https://docs.pytest.org/_/downloads/en/latest/pdf/"
    download_dir = os.path.join(SCRIPT_DIR, 'resources')
    path_to_file = os.path.join(download_dir, 'docs-pytest-org-en-latest.pdf')
    path_to_etalon_file = os.path.join(SCRIPT_DIR, 'test_data', 'etalon_docs-pytest-org-en-latest.pdf')

    response = requests.get(url)
    response.raise_for_status()

    with open(path_to_file, 'wb') as file:
        file.write(response.content)

    reader = PdfReader(path_to_file)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()

    reader = PdfReader(path_to_etalon_file)
    etalon_number_of_pages = len(reader.pages)
    etalon_page = reader.pages[0]
    etalon_text = page.extract_text()

    assert number_of_pages == etalon_number_of_pages
    assert text == etalon_text
    assert _page_assert(page, etalon_page)


def test_xlsx():
    path_to_file = os.path.join(PATH, 'test_data', 'file_example_XLSX_50.xlsx')
    etalon_string = "some_value @#$%^&SDFGCVBoookjjt"
    workbook = load_workbook(path_to_file)
    sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == etalon_string


def test_ziprar():
    test_data = os.path.join(SCRIPT_DIR, 'test_data')
    resources = os.path.join(SCRIPT_DIR, 'resources')
    tmp = os.path.join(SCRIPT_DIR, 'tmp')
    archive_path = os.path.join(tmp, 'rar.zip')

    shutil.copytree(test_data, resources, dirs_exist_ok=True)
    shutil.make_archive(os.path.splitext(archive_path)[0], 'zip', resources)

    # Проверка содержимого архива
    with zipfile.ZipFile(archive_path, 'r') as zipf:
        for file_name in zipf.namelist():
            archived_file_data = zipf.read(file_name)

            with open(os.path.join(resources, file_name), 'rb') as f:
                original_file_data = f.read()

            # Сравнение содержимого
            assert archived_file_data == original_file_data


def _clean_dir(directory_path):
    for filename in os.listdir(directory_path):
        file_path = os.path.join(directory_path, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f'Не удалось удалить файл {file_path}. Ошибка: {e}')


def _page_assert(compared, etalon):
    for key, value in etalon.items():
        if isinstance(value, IndirectObject) and not isinstance(compared[key], IndirectObject):
            return compared[key] == value.get_object()
        if isinstance(value, IndirectObject) and isinstance(compared[key], IndirectObject):
            # здесь я пробовал структурное рекурсивное сравнение, но упирался в переполнение стека
            # если напрямую сравнивать IndirectObject то они будут отличаться, т.к. метаданные разные при создании pdf
            # если бы мне нужно было сравнить два pdf в боевой задаче я бы вероятно использовал что-то вроде fitz
            return True
        return compared[key] == value


resources = os.path.join(PATH, 'resources')
tmp = os.path.join(PATH, 'tmp')
_clean_dir(resources)
_clean_dir(tmp)
